"""
Exportador de reportes a Excel (.xlsx) con openpyxl.
"""
import io
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from app.services.financiero import obtener_matriz_deudas
from app.models.pago import Pago
from app.models.apartamento import Apartamento
from app.config import get_settings

settings = get_settings()

COLOR_VERDE = "2D5A43"
COLOR_ROJO = "C0392B"
COLOR_AMARILLO = "E67E22"


def _estilo_encabezado(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=COLOR_VERDE)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _ajustar_columnas(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def generar_excel_reporte(db: Session, periodo: str) -> bytes:
    """
    Genera un Excel con 3 hojas:
      1. Matriz de Deudas
      2. Pagos Aprobados del período
      3. Saldos a Favor
    """
    wb = Workbook()

    # ── Hoja 1: Matriz de Deudas ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "Matriz de Deudas"

    encabezados1 = [
        "Apto", "Piso", "Torre", "Propietario",
        "Teléfono", "Email", "Deuda (USD)", "Deuda (VES)",
        "Saldo a Favor (USD)", "Meses Adeudados", "Estado",
    ]
    ws1.append(encabezados1)
    for cell in ws1[1]:
        _estilo_encabezado(cell)
    ws1.row_dimensions[1].height = 20

    matriz = obtener_matriz_deudas(db)
    for fila in matriz:
        estado = fila["estado"]
        row = [
            fila["numero_apto"], fila["piso"] or "-", fila["torre"],
            fila["propietario"], fila["telefono"], fila["email"],
            fila["deuda_total_usd"], fila["deuda_total_ves"],
            fila["saldo_favor_usd"],
            ", ".join(fila["meses_adeudados"]) or "Ninguno",
            estado.upper(),
        ]
        ws1.append(row)

    _ajustar_columnas(ws1)

    # ── Hoja 2: Pagos Aprobados ───────────────────────────────────
    ws2 = wb.create_sheet("Pagos Aprobados")
    encabezados2 = [
        "Apto", "Propietario", "Período", "Método de Pago",
        "Referencia", "Banco", "Monto Declarado", "Moneda",
        "Equivalente USD", "Tasa BCV", "Fecha Aprobación",
    ]
    ws2.append(encabezados2)
    for cell in ws2[1]:
        _estilo_encabezado(cell)

    pagos = (
        db.query(Pago)
        .join(Pago.recibo)
        .filter(
            Pago.estado_conciliacion == "aprobado",
        )
        .all()
    )
    for pago in pagos:
        apto = pago.apartamento
        propietario = apto.propietario
        ws2.append([
            apto.numero_apto,
            f"{propietario.nombre} {propietario.apellido}" if propietario else "-",
            pago.recibo.mes_periodo if pago.recibo else "-",
            pago.metodo_pago,
            pago.referencia_bancaria,
            pago.banco_origen or "-",
            float(pago.monto_declarado),
            pago.moneda_pago,
            float(pago.monto_equivalente_usd),
            float(pago.tasa_bcv_aplicada),
            pago.fecha_aprobacion.strftime("%d/%m/%Y %H:%M") if pago.fecha_aprobacion else "-",
        ])

    _ajustar_columnas(ws2)

    # ── Hoja 3: Saldos a Favor ────────────────────────────────────
    ws3 = wb.create_sheet("Saldos a Favor")
    encabezados3 = ["Apto", "Piso", "Propietario", "Teléfono", "Saldo a Favor (USD)"]
    ws3.append(encabezados3)
    for cell in ws3[1]:
        _estilo_encabezado(cell)

    aptos_con_saldo = (
        db.query(Apartamento)
        .filter(Apartamento.saldo_favor_usd > 0)
        .all()
    )
    for apto in aptos_con_saldo:
        p = apto.propietario
        ws3.append([
            apto.numero_apto,
            apto.piso or "-",
            f"{p.nombre} {p.apellido}" if p else "-",
            p.telefono_whatsapp if p else "-",
            float(apto.saldo_favor_usd),
        ])

    _ajustar_columnas(ws3)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
